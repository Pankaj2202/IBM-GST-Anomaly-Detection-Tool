"""
Build the GST Anomaly Detection Excel report (the showcase deliverable).
Multi-sheet: Read Me | Executive Dashboard | Model Performance |
High-Risk Review Queue | Rule-Flag Summary | Scored Data.
"""

import json
import joblib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import BarChart, Reference, PieChart

# ---- palette (JK / exec-report style) -------------------------------------
NAVY = "1F3A5F"; STEEL = "2E5E8C"; LIGHT = "EAF0F7"; GREY = "F4F6F9"
RED = "C0392B"; AMBER = "E67E22"; GREEN = "27AE60"
RED_F = "F8D7DA"; AMBER_F = "FDEBD0"; GREEN_F = "D5F5E3"
WHITE = "FFFFFF"; DARK = "1A1A1A"
FONT = "Arial"

thin = Side(style="thin", color="D0D7E2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, cell, text, size=11, color=WHITE, fill=NAVY, align="left"):
    c = ws[cell]; c.value = text
    c.font = Font(name=FONT, bold=True, size=size, color=color)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border


def cell(ws, ref, val, bold=False, size=10, color=DARK, fill=None,
         align="left", num=None, border_on=True):
    c = ws[ref]; c.value = val
    c.font = Font(name=FONT, bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if num:
        c.number_format = num
    if border_on:
        c.border = border
    return c


def title_block(ws, title, subtitle):
    ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")
    cell(ws, "A1", title, bold=True, size=16, color=WHITE, fill=NAVY,
         align="left", border_on=False)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell(ws, "A2", subtitle, size=9, color=WHITE, fill=STEEL,
         align="left", border_on=False)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18


# ===========================================================================
def main():
    scored = pd.read_pickle("scored_population.pkl")
    metrics = json.load(open("model_metrics.json"))
    meta = joblib.load("gst_rf_model.joblib")["ref_meta"]

    wb = Workbook()

    # ---------------------------------------------------------------- Read Me
    ws = wb.active; ws.title = "Read Me"
    ws.sheet_view.showGridLines = False
    title_block(ws, "GST Anomaly Detection Tool",
                "JK Cement  |  Supply Chain Accounts Payable  |  Fraud-Risk Scoring Engine")
    rows = [
        ("", ""),
        ("PURPOSE", ""),
        ("", "Flags GST invoices with a high risk of error or fraudulent ITC and ranks them for "
             "human review, replacing slow manual checking. It does NOT auto-reject invoices."),
        ("", ""),
        ("HOW IT WORKS", ""),
        ("1. Rule engine", "Deterministic, auditable checks on every invoice - the 5 validation "
                           "algorithms + GST checkpoints. Each flag is fully explainable."),
        ("2. ML model", "A Random Forest learns the combined pattern of historically-flagged "
                        "anomalies and outputs a single Risk Score (0-1) per invoice."),
        ("3. Risk band", "Score is bucketed LOW / MEDIUM / HIGH so reviewers work the queue "
                         "top-down by priority."),
        ("", ""),
        ("VALIDATION ALGORITHMS", ""),
        ("R1 Invoice No.", "Present & vendor exists in the active vendor master."),
        ("R2 Invoice Date", "Cannot be a future date vs posting date."),
        ("R3 Invoice Amount", "Deviation from the vendor's own historical median (mean/median %)."),
        ("R4 Tax Code", "Valid code, consistent with Inter/Intra (Ship-To vs Bill-To)."),
        ("R5 Business Place", "Valid & maps to a known JK Cement recipient GSTIN."),
        ("R6 Filing Delay", "Stale / old invoices beyond the ageing threshold."),
        ("", ""),
        ("IMPORTANT - READ", ""),
        ("Risk, not verdict", "A HIGH score means 'review this first', NOT 'this is fraud'. The "
                             "labels are anomalies/exceptions, so every flagged item still needs "
                             "human confirmation before any action against a vendor."),
        ("", ""),
        ("DAILY USE", ""),
        ("Command", "python score_daily.py  <posted_invoices.xlsx>   ->  daily_review_queue.xlsx"),
        ("Retrain", "python train.py   (rebuilds the model from ACL + ARM history)"),
        ("", ""),
        ("RISK BAND LEGEND", ""),
        ("HIGH", "Score >= 0.70  -  review first"),
        ("MEDIUM", "Score 0.40 - 0.70  -  review next"),
        ("LOW", "Score < 0.40  -  routine"),
    ]
    r = 4
    for a, b in rows:
        if a and not b:  # section header
            ws.merge_cells(f"A{r}:H{r}")
            cell(ws, f"A{r}", a, bold=True, size=11, color=WHITE, fill=STEEL, border_on=False)
        else:
            cell(ws, f"A{r}", a, bold=True, size=10, color=NAVY, border_on=False)
            ws.merge_cells(f"B{r}:H{r}")
            cell(ws, f"B{r}", b, size=10, color=DARK, border_on=False)
            ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if b and len(b) > 70:
            ws.row_dimensions[r].height = 30
        r += 1
    # band legend colors
    for rr, fill in [(r-3, RED_F), (r-2, AMBER_F), (r-1, GREEN_F)]:
        ws[f"A{rr}"].fill = PatternFill("solid", fgColor=fill)
    ws.column_dimensions["A"].width = 20
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14

    # ------------------------------------------------------ Executive Dashboard
    build_dashboard(wb, scored, metrics, meta)
    build_model_perf(wb, metrics)
    build_review_queue(wb, scored)
    build_rule_summary(wb, scored)
    build_scored_data(wb, scored)

    wb.save("GST_Anomaly_Detection_Report.xlsx")
    print("Saved GST_Anomaly_Detection_Report.xlsx")


def build_dashboard(wb, scored, metrics, meta):
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Executive Dashboard",
                "Population scoring summary  |  model trained on FY24-25 ACL + ARM history")

    n = len(scored)
    anom = int(scored["is_anomaly"].sum()) if "is_anomaly" in scored else None
    hi = int((scored["risk_band"] == "HIGH").sum())
    md = int((scored["risk_band"] == "MEDIUM").sum())
    lo = n - hi - md

    # KPI cards (row 4-6)
    kpis = [
        ("Invoices Scored", f"{n:,}", NAVY),
        ("Historical Anomalies", f"{anom:,}" if anom is not None else "-", STEEL),
        ("HIGH-Risk Queue", f"{hi:,}", RED),
        ("Model Recall", f"{metrics['recall']*100:.1f}%", GREEN),
    ]
    col = 1
    for label, val, color in kpis:
        cl = get_column_letter(col); cr = get_column_letter(col+1)
        ws.merge_cells(f"{cl}4:{cr}4"); ws.merge_cells(f"{cl}5:{cr}5")
        cell(ws, f"{cl}4", val, bold=True, size=20, color=WHITE, fill=color,
             align="center", border_on=False)
        cell(ws, f"{cl}5", label, bold=True, size=9, color=WHITE, fill=color,
             align="center", border_on=False)
        col += 2
    ws.row_dimensions[4].height = 34

    # Risk band distribution table
    hdr(ws, "A8", "Risk Band"); hdr(ws, "B8", "Count"); hdr(ws, "C8", "% of Pop", align="right")
    band_rows = [("HIGH", hi, RED_F), ("MEDIUM", md, AMBER_F), ("LOW", lo, GREEN_F)]
    rr = 9
    for name, cnt, fill in band_rows:
        cell(ws, f"A{rr}", name, bold=True, fill=fill)
        cell(ws, f"B{rr}", cnt, align="right", num="#,##0")
        cell(ws, f"C{rr}", f"=B{rr}/B12", align="right", num="0.0%")
        rr += 1
    cell(ws, "A12", "TOTAL", bold=True, fill=LIGHT)
    cell(ws, "B12", "=SUM(B9:B11)", bold=True, align="right", num="#,##0", fill=LIGHT)
    cell(ws, "C12", "=B12/B12", bold=True, align="right", num="0.0%", fill=LIGHT)

    # anomaly type breakdown
    hdr(ws, "E8", "Anomaly Type (historical)"); hdr(ws, "F8", "Count", align="right")
    if "anomaly_type" in scored:
        at = scored[scored["anomaly_type"] != "None"]["anomaly_type"].value_counts()
        rr = 9
        for name, cnt in at.items():
            cell(ws, f"E{rr}", name); cell(ws, f"F{rr}", int(cnt), align="right", num="#,##0")
            rr += 1
        cell(ws, f"E{rr}", "TOTAL", bold=True, fill=LIGHT)
        cell(ws, f"F{rr}", f"=SUM(F9:F{rr-1})", bold=True, align="right", num="#,##0", fill=LIGHT)

    # pie chart for risk bands
    pie = PieChart(); pie.title = "Risk Band Mix"
    data = Reference(ws, min_col=2, min_row=9, max_row=11)
    cats = Reference(ws, min_col=1, min_row=9, max_row=11)
    pie.add_data(data); pie.set_categories(cats); pie.height = 6.5; pie.width = 9
    ws.add_chart(pie, "A14")

    # bar chart for anomaly types
    if "anomaly_type" in scored:
        bar = BarChart(); bar.title = "Anomalies by Type"; bar.type = "bar"
        bar.legend = None
        dref = Reference(ws, min_col=6, min_row=9, max_row=9+len(at)-1)
        cref = Reference(ws, min_col=5, min_row=9, max_row=9+len(at)-1)
        bar.add_data(dref); bar.set_categories(cref); bar.height = 6.5; bar.width = 11
        ws.add_chart(bar, "E14")

    ws.column_dimensions["A"].width = 16
    for c in "BCD":
        ws.column_dimensions[c].width = 12
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 12


def build_model_perf(wb, metrics):
    ws = wb.create_sheet("Model Performance")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Model Performance",
                "Random Forest  |  300 trees  |  class-weight balanced  |  25% hold-out test")

    hdr(ws, "A4", "Metric"); hdr(ws, "B4", "Value", align="right")
    mrows = [
        ("Accuracy", metrics["accuracy"], "0.0%"),
        ("Precision", metrics["precision"], "0.0%"),
        ("Recall (fraud catch-rate)", metrics["recall"], "0.0%"),
        ("F1 Score", metrics["f1"], "0.000"),
        ("ROC-AUC", metrics["roc_auc"], "0.000"),
        ("Train rows", metrics["n_train"], "#,##0"),
        ("Test rows", metrics["n_test"], "#,##0"),
    ]
    r = 5
    for name, val, fmt in mrows:
        cell(ws, f"A{r}", name, bold=(name == "Recall (fraud catch-rate)"))
        cell(ws, f"B{r}", val, align="right", num=fmt,
             bold=(name == "Recall (fraud catch-rate)"),
             fill=GREEN_F if name == "Recall (fraud catch-rate)" else None)
        r += 1

    # confusion matrix
    cm = metrics["confusion_matrix"]
    hdr(ws, "D4", "Confusion Matrix"); ws.merge_cells("D4:F4")
    cell(ws, "E5", "Pred Normal", bold=True, fill=LIGHT, align="center")
    cell(ws, "F5", "Pred Anomaly", bold=True, fill=LIGHT, align="center")
    cell(ws, "D6", "Actual Normal", bold=True, fill=LIGHT)
    cell(ws, "D7", "Actual Anomaly", bold=True, fill=LIGHT)
    cell(ws, "E6", cm[0][0], align="center", num="#,##0", fill=GREEN_F)
    cell(ws, "F6", cm[0][1], align="center", num="#,##0", fill=AMBER_F)
    cell(ws, "E7", cm[1][0], align="center", num="#,##0", fill=RED_F)
    cell(ws, "F7", cm[1][1], align="center", num="#,##0", fill=GREEN_F)
    cell(ws, "D9", "TP = caught anomalies | FN = missed | FP = false alarms",
         size=8, color="555555", border_on=False)

    # feature importance
    hdr(ws, "A14", "Feature"); hdr(ws, "B14", "Importance", align="right")
    fi = metrics["feature_importance"]
    r = 15
    for name, imp in fi:
        cell(ws, f"A{r}", name); cell(ws, f"B{r}", imp, align="right", num="0.000")
        r += 1
    ws.conditional_formatting.add(
        f"B15:B{r-1}",
        DataBarRule(start_type="num", start_value=0, end_type="num",
                    end_value=max(i for _, i in fi), color=STEEL))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    for c in "DEF":
        ws.column_dimensions[c].width = 15


def _review_reason(row):
    """Human-readable 'why flagged' - rule reasons, else a model-derived note."""
    rr = row.get("rule_reasons") or ""
    if isinstance(rr, str) and rr.strip():
        return rr
    notes = []
    if row.get("amount_dev_pct", 0) and row["amount_dev_pct"] > 20:
        notes.append(f"Amount {row['amount_dev_pct']:.0f}% off vendor norm")
    if abs(row.get("doc_amount", 0) or 0) >= 1_000_000:
        notes.append("High-value invoice")
    if row.get("filing_delay_days", 0) and row["filing_delay_days"] > 90:
        notes.append(f"{int(row['filing_delay_days'])}d filing delay")
    if not notes:
        notes.append("ML risk pattern (vendor/amount profile)")
    return "ML: " + "; ".join(notes)


def build_review_queue(wb, scored):
    ws = wb.create_sheet("High-Risk Review Queue")
    ws.sheet_view.showGridLines = False
    title_block(ws, "High-Risk Review Queue",
                "HIGH-risk invoices, ranked by score - work this list top-down")

    q = scored[scored["risk_band"] == "HIGH"].sort_values(
        "risk_score", ascending=False).copy()
    q["review_reason"] = q.apply(_review_reason, axis=1)
    cols = [("document_no", "SAP Doc No", 14, "#,##0"),
            ("vendor_code", "Vendor Code", 12, None),
            ("vendor_name", "Vendor Name", 30, None),
            ("invoice_no", "Invoice No", 18, None),
            ("doc_date", "Invoice Date", 12, "dd-mmm-yy"),
            ("pstng_date", "Posting Date", 12, "dd-mmm-yy"),
            ("doc_amount", "Amount (INR)", 14, "#,##0;(#,##0)"),
            ("filing_delay_days", "Delay (d)", 9, "0"),
            ("amount_dev_pct", "Amt Dev %", 10, "0.0"),
            ("rule_flag_count", "Flags", 7, "0"),
            ("risk_score", "Risk Score", 11, "0.000"),
            ("review_reason", "Reasons", 46, None)]

    hrow = 4
    for j, (_, label, w, _) in enumerate(cols, start=1):
        hdr(ws, f"{get_column_letter(j)}{hrow}", label,
            align="right" if label in ("Amount (INR)", "Risk Score", "Amt Dev %") else "left")
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{hrow+1}"

    rr = hrow + 1
    for _, row in q.iterrows():
        for j, (key, _, _, fmt) in enumerate(cols, start=1):
            v = row.get(key)
            if pd.isna(v):
                v = ""
            if key in ("doc_date", "pstng_date") and v != "" and not isinstance(v, str):
                v = pd.Timestamp(v).to_pydatetime()
            align = "right" if fmt and ("#" in fmt or fmt in ("0.000", "0.0", "0")) else "left"
            cell(ws, f"{get_column_letter(j)}{rr}", v, num=fmt, align=align, size=9)
        rr += 1

    # color-scale the risk score column + bold red band
    score_col = get_column_letter(11)
    ws.conditional_formatting.add(
        f"{score_col}{hrow+1}:{score_col}{rr-1}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.85"],
                   fill=PatternFill("solid", fgColor=RED_F),
                   font=Font(name=FONT, bold=True, color=RED, size=9)))
    cell(ws, f"A{rr+1}", f"{len(q):,} HIGH-risk invoices listed.",
         bold=True, color=NAVY, border_on=False)


def build_rule_summary(wb, scored):
    ws = wb.create_sheet("Rule-Flag Summary")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Rule-Flag Summary",
                "How often each deterministic validation rule fired across the population")

    flag_labels = {
        "flag_future_date": "R2  Future-dated invoice",
        "flag_old_invoice": "R6  Stale / old invoice",
        "flag_invoice_no_missing": "R1  Invoice number missing",
        "flag_vendor_not_in_master": "R1  Vendor not in active master",
        "flag_amount_outlier": "R3  Amount deviates from vendor norm",
        "flag_business_place_invalid": "R5  Invalid business place",
        "flag_tax_code_invalid": "R4  Invalid tax code",
        "flag_round_amount": "R7  Large round-number amount",
    }
    hdr(ws, "A4", "Validation Rule"); hdr(ws, "B4", "Invoices Flagged", align="right")
    hdr(ws, "C4", "% of Pop", align="right")
    n = len(scored)
    r = 5
    for key, label in flag_labels.items():
        cnt = int(scored[key].sum()) if key in scored else 0
        cell(ws, f"A{r}", label)
        cell(ws, f"B{r}", cnt, align="right", num="#,##0")
        cell(ws, f"C{r}", cnt / n, align="right", num="0.0%")
        r += 1
    ws.conditional_formatting.add(
        f"B5:B{r-1}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=AMBER))
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12


def build_scored_data(wb, scored):
    ws = wb.create_sheet("Scored Data (All)")
    cols = ["document_no", "vendor_code", "vendor_name", "invoice_no",
            "doc_date", "pstng_date", "doc_amount", "filing_delay_days",
            "amount_dev_pct", "rule_flag_count", "rule_reasons",
            "risk_score", "risk_band", "anomaly_type"]
    cols = [c for c in cols if c in scored.columns]
    out = scored[cols].copy()
    out.columns = [c.replace("_", " ").title() for c in cols]
    for j, name in enumerate(out.columns, start=1):
        hdr(ws, f"{get_column_letter(j)}1", name)
        ws.column_dimensions[get_column_letter(j)].width = 16
    for i, (_, row) in enumerate(out.iterrows(), start=2):
        for j, name in enumerate(out.columns, start=1):
            v = row[name]
            if pd.isna(v):
                v = ""
            ws.cell(row=i, column=j, value=v).font = Font(name=FONT, size=8)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(out.columns))}{len(out)+1}"


if __name__ == "__main__":
    main()
